import timeit
from flask import Flask, request, jsonify, send_from_directory, render_template
from openpyxl import load_workbook, Workbook
import os
import requests
import qdrant_client
from dotenv import load_dotenv
from flask_cors import CORS
from werkzeug.utils import secure_filename

import cloudinary
import cloudinary.uploader
from cloudinary.utils import cloudinary_url

# Import llama_index components
from llama_index.core import VectorStoreIndex, get_response_synthesizer
from llama_index.core.node_parser import SentenceSplitter
from llama_index.core.retrievers import VectorIndexRetriever
from llama_index.core import Document
from llama_index.vector_stores.qdrant import QdrantVectorStore
from llama_index.embeddings.openai import OpenAIEmbedding
from llama_index.core.ingestion import IngestionPipeline
from llama_index.core.response.pprint_utils import pprint_response
from llama_index.core.query_engine import CustomQueryEngine
from llama_index.core.retrievers import BaseRetriever
from llama_index.core.response_synthesizers import BaseSynthesizer
from llama_index.llms.openai import OpenAI
from llama_index.core import PromptTemplate

# Initialize Flask application
app = Flask(__name__)
CORS(app)

# Configuration       
cloudinary.config( 
    cloud_name = "dtjivws2c", 
    api_key = "897949739888447", 
    api_secret = "3fFPMLX_uRSUXR2noV-SmQXFP-s", # Click 'View Credentials' below to copy your API secret
)

def extractor(file_path):
    wb = load_workbook(filename=file_path, read_only=True)
    ws = wb.active

    data = []
    header = [cell.value for cell in next(ws.iter_rows())]
    col_index = {name: index for index, name in enumerate(header)}
    keys = ['Title', 'Abstract', 'English description', 'Claims']
    available_cols = {key: col_index.get(key) for key in keys if key in col_index}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(row):
            row_dict = {}
            for key, index in available_cols.items():
                row_dict[key.lower()] = row[index] if index is not None else None
            data.append(row_dict)
    return data

def newFileSaver(relevancy, file_path):
    workbook = load_workbook(filename=file_path)
    sheet = workbook.active
    relevancy_header = 'Relevancy predicted'
    comments_header = 'Comments made'

    empty_column = None
    for cell in sheet[1]:
        if cell.value is None:
            empty_column = cell.column
            break

    if empty_column is None:
        empty_column = sheet.max_column + 1

    sheet.cell(row=1, column=empty_column, value=relevancy_header)
    sheet.cell(row=1, column=empty_column + 1, value=comments_header)

    for i, (status, comment) in enumerate(relevancy, start=2):
        sheet.cell(row=i, column=empty_column, value=status)
        sheet.cell(row=i, column=empty_column + 1, value=comment)

    workbook.save(filename=file_path)

    new_workbook = Workbook()
    new_sheet = new_workbook.active
    for col_num, cell in enumerate(sheet[1], 1):
        new_sheet.cell(row=1, column=col_num, value=cell.value)

    new_row_idx = 2
    for row in sheet.iter_rows(min_row=2, values_only=False):
        if row[empty_column-1].value == 'R':
            for col_num, cell in enumerate(row, 1):
                new_sheet.cell(row=new_row_idx, column=col_num, value=cell.value)
            new_row_idx += 1

    new_file_path = os.path.splitext(file_path)[0] + '_relevant_only.xlsx'
    new_workbook.save(filename=new_file_path)

    return file_path, new_file_path

    
from openpyxl import load_workbook
from openpyxl import Workbook
import os

def newFileSaver1(category_data, file_path, categories):
    # Load the existing workbook
    workbook = load_workbook(filename=file_path)
    sheet = workbook.active

    # Find the first empty column in the first row
    empty_column = None
    for cell in sheet[1]:
        if cell.value is None:
            empty_column = cell.column
            break

    # If no empty column is found, set the empty_column to the next column
    if empty_column is None:
        empty_column = sheet.max_column + 1

    # Write the category headers in the first row
    for i, category in enumerate(categories):
        sheet.cell(row=1, column=empty_column+i, value=category)

    # Write the category data to the workbook starting from the second row
    for i, data_row in enumerate(category_data, start=2):
        for j, data in enumerate(data_row):
            sheet.cell(row=i, column=empty_column+j, value=data)

    # Save the workbook
    workbook.save(filename=file_path)

    # Create a new file name by appending '_category_predicted' to the original file name
    base, ext = os.path.splitext(file_path)
    new_file_path = f"{base}_category_predicted{ext}"

    # Save a copy of the workbook with the new file name
    workbook.save(filename=new_file_path)

    # Return the file path of the new workbook
    return new_file_path


def callLLM(rowData,catNquery):

    # Create documents from the dictionary item
    documents = [Document(text=f"{key}: {val}") for key, val in rowData.items()]

    client = qdrant_client.QdrantClient(location=":memory:")
    vector_store = QdrantVectorStore(client=client, collection_name="test_store")

    # Ingest documents into a vector store
    pipeline = IngestionPipeline(
        transformations=[
            SentenceSplitter(chunk_size=128, chunk_overlap=5),
            OpenAIEmbedding(),
        ],
        vector_store=vector_store,
    )
    pipeline.run(documents=documents)
    index = VectorStoreIndex.from_vector_store(vector_store)

    retriever = VectorIndexRetriever(index=index, similarity_top_k=5)
    response_synthesizer = get_response_synthesizer(response_mode="tree_summarize")
    llm = OpenAI(model="gpt-4o")

    class RAGStringQueryEngine(CustomQueryEngine):
        """
        Custom Query Engine for RAG (Retrieval-Augmented Generation).
        """
        retriever: BaseRetriever
        response_synthesizer: BaseSynthesizer
        llm: OpenAI
        qa_prompt: PromptTemplate

        def custom_query(self, query_str: str):
            """
            Perform a custom query.
            
            Args:
                query_str (str): The query string.

            Returns:
                str: The response from the query.
            """
            nodes = self.retriever.retrieve(query_str)
            context_str = "\n\n".join([n.node.get_content() for n in nodes])

            response = self.llm.complete(qa_prompt.format(context_str=context_str, query_str=query_str))
            return str(response)

    

    rowCatData = []


    for column in catNquery:

        # note time for each loop
        start = timeit.default_timer()


        category =  column[0] 
        query = column[1]
        qa_prompt = PromptTemplate(
            "Context information is"
            " below.\n---------------------\n{context_str}\n---------------------\nUsing"
            " only the context information and not your own knowledge, answer"
            " the question: {query_str}\n"
            " Only give to-the-point answers, try to minimise the output length while also providing all the neccessary information. \n"
        )
        query_engine = RAGStringQueryEngine(
        retriever=retriever,
        response_synthesizer=response_synthesizer,
        llm=llm,
        qa_prompt=qa_prompt,
        )
        response = query_engine.query(query+"Name of the category is: "+category)
        pprint_response(response)
        rowCatData.append(str(response))
        print(f"Time taken for {category} is {timeit.default_timer() - start}")
        print("--"*50)




    return rowCatData

def backend1(datalist, catNquery, file_path):

    load_dotenv()  # Load environment variables

    llama_api_key = os.getenv('LLAMA_CLOUD_API_KEY')
    if llama_api_key is None:
        raise ValueError("LLAMA_CLOUD_API_KEY not found in environment variables")

    categories = [cat[0] for cat in catNquery]

    category_data = []

    for rowData in datalist:
        # Call the LLM on the current rowData
        result = callLLM(rowData, catNquery)
        
        # Append the categories result to the category data list
        category_data.append(result)

    outputFilePath = newFileSaver1(category_data, file_path, categories)

    return outputFilePath

def extract_reason(text):
    parts = text.split("Reason: ", 1)
    return parts[1] if len(parts) > 1 else ""

def extract_related(text):
    return '1R1' in text

def backend(dict_item, user_query):
    load_dotenv()
    llama_api_key = os.getenv('LLAMA_CLOUD_API_KEY')
    if llama_api_key is None:
        raise ValueError("LLAMA_CLOUD_API_KEY not found in environment variables")

    documents = [Document(text=f"{key}: {val}") for key, val in dict_item.items()]
    client = qdrant_client.QdrantClient(location=":memory:")
    vector_store = QdrantVectorStore(client=client, collection_name="test_store")

    pipeline = IngestionPipeline(
        transformations=[
            SentenceSplitter(chunk_size=128, chunk_overlap=5),
            OpenAIEmbedding(),
        ],
        vector_store=vector_store,
    )
    pipeline.run(documents=documents)
    index = VectorStoreIndex.from_vector_store(vector_store)

    class RAGStringQueryEngine(CustomQueryEngine):
        retriever: BaseRetriever
        response_synthesizer: BaseSynthesizer
        llm: OpenAI
        qa_prompt: PromptTemplate

        def custom_query(self, query_str: str):
            nodes = self.retriever.retrieve(query_str)
            context_str = "\n\n".join([n.node.get_content() for n in nodes])
            response = self.llm.complete(self.qa_prompt.format(context_str=context_str, query_str=query_str))
            return str(response)

    qa_prompt = PromptTemplate(
        "You are an AI assistant that predicts relevancy of a 'Document' with a certain 'Statement'. If it is Relevant then return output as '1R1', otherwise '0R0'. If output is '1R1', then state the 'Reason' which makes it relevant with the help of information present in 'Document'. \n"
        "For example 1:\n"
        "Document:" + '''title: Composition, application of the composition, cosmetic preparation hydrogel bio-mask in the form of a compress, method of manufacturing the preparation
        Background of the invention.
        hydrogel bio-mask composed of natural materials and active ingredients, designed for cosmetic applications to enhance skin health. The hydrogel matrix provides a natural and effective medium for delivering active ingredients to the skin. the composition of the hydrogel bio-mask and its natural active ingredients. The following are the key points regarding the specific ingredients mentioned
        Hydrogel Matrix: The document emphasizes the use of a hydrogel matrix obtained from natural sources. Natural Active Ingredients: The hydrogel bio-mask includes various natural active ingredients intended for cosmetic use.''' + "\n"
        "Statement:" + "Mannuronic acid or avocado or alginate from algae should be used in a cosmetic formulation for any skin claim" + "\n"
        "Output: '0R0' \n"
        "Reason:  It is not mentioning the use of Mannuronic acid, alginate, or avocado but having skin claim for cosmetics \n"
        "For example 2:\n"
        "Document:" + '''the use of mannuronic acid derivatives and alginate from algae in cosmetic formulations aimed at improving skin health by providing anti-photoaging benefits, moisture retention, antioxidant protection, and enzyme inhibition. The derivatives form an invisible film on the skin, protecting against UV damage and maintaining a moist environment. They exhibit strong antioxidant capabilities and inhibit enzymes like tyrosinase and elastase, reducing melanin production and collagen degradation.
        The primary focus of the patent is on alginate oligosaccharide derivatives derived from brown algae. These are used for their moisture absorption, antioxidation, and enzyme inhibition properties in skincare products.''' + "\n"
        "Statement:" + "Mannuronic acid or avocado or alginate from algae should be used in a cosmetic formulation for any skin claim" + "\n"
        "Output: '1R1' \n"
        "Reason :  Mannuronic acid and alginate from algae is used for different skin claims in a cosmetic product \n"
        "For example 3:\n"
        "Document:" + '''title: Use of brown algae water extract for preparing blue light resistant skin external product
        Background of the invention.
        using brown algae extract containing fucoidan for preparing topical skin care products that protect against blue light exposure. These products aim to improve skin health by reducing wrinkles and enhancing brightness, particularly for individuals frequently exposed to blue light. The invention emphasizes the benefits of fucoidan in long-term skin care.
        The present invention provides a use of a brown algae extract for preparing a skin topical product for anti-blue light, wherein the product is provided to a subject exposed to blue light, and the brown algae extract contains fucoidan.''' + "\n"
        "Statement:" + "Mannuronic acid or avocado or alginate from algae should be used in a cosmetic formulation for any skin claim" + "\n"
        "Output: '1R1' \n"
        "Reason: Alginate from Brown Algae is used for protecting against blue light in skincare products \n"
        "Using the below given Document and Statement, provide the Output and Reason\n"
        "Document: {context_str}\n"
        "Statement: {query_str}\n"
        "Output: \n"
        "Reason: "
    )

    retriever = VectorIndexRetriever(index=index, similarity_top_k=5)
    response_synthesizer = get_response_synthesizer(response_mode="tree_summarize")
    llm = OpenAI(model="gpt-4o")

    query_engine = RAGStringQueryEngine(
        retriever=retriever,
        response_synthesizer=response_synthesizer,
        llm=llm,
        qa_prompt=qa_prompt,
    )
    response = query_engine.query(user_query)
    pprint_response(response)

    response_str = str(response)
    return extract_related(response_str), extract_reason(response_str)


@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({'error': 'No file part'})
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No selected file'})

    if file:
        upload_options ={
            'resource_type': 'raw',
            'filename_override': 'uploads.xlsx'
        }

        filename = secure_filename(file.filename)
        # Upload the file to Cloudinary
        upload_result = cloudinary.uploader.upload(file, **upload_options)
        
        # Get the URL of the uploaded file
        download_url = upload_result.get("url")
        print(download_url)
        # file.save(os.path.join('./uploads', filename))
        # file_path = os.path.join('./uploads', filename)
        return jsonify({'file_path': download_url})

@app.route('/process', methods=['POST'])
def process_file():
    data = request.json
    user_query = data.get('user_query')
    file_path = data.get('file_path')
    response = requests.get(file_path)
    # response.raise_for_status()
    
    # Save the downloaded file to a temporary location
    temp_filename = os.path.join('./uploads', 'uploaded_file.xlsx')
    with open(temp_filename, 'wb') as temp_file:
        temp_file.write(response.content)

    # if not file_path:
    #     return jsonify({'error': 'File path is missing'})

    datalist = extractor(temp_filename)

    if not datalist:
        return jsonify({'error': 'Failed to extract data from the file'})

    relevancy = []
    for dict_item in datalist:
        result = backend(dict_item, user_query)
        status = "R" if result[0] else "NR"
        relevancy.append((status, result[1]))

    outputFilePath, newFilePath = newFileSaver(relevancy, temp_filename)
    return jsonify({'Path': os.path.basename(outputFilePath), 'FilteredPath': os.path.basename(newFilePath)})

@app.route('/download/<path:filename>', methods=['GET'])
def download(filename):
    directory = './uploads'
    return send_from_directory(directory, filename, as_attachment=True)

@app.route('/download-category/<path:filename>', methods=['GET'])
def download_category(filename):
    directory ='./uploads'
    return send_from_directory(directory,filename, as_attachment =True )
@app.route('/process-relevant-only', methods =['POST'] )
def process_relevant_file():

    data = request.json
    print(data)
    
    file_path = data.get('relevant_only_file_path')
    filename = os.path.join('./uploads', 'uploaded_file_relevant_only.xlsx')
    print(filename)
    datalist = extractor(filename)
    catNquery = data.get('search_terms')
    print(catNquery)

    if datalist[-1]['title'] is None:
        datalist.pop()

    outputFilePath1 = backend1(datalist,catNquery,filename)
    # outputFilePath = file_path
    print(os.path.basename(outputFilePath1))
    base_filename = os.path.basename(outputFilePath1)
    
    return jsonify({'Path': base_filename})

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
